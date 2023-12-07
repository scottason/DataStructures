from .Books import Book
import openpyxl

class LibraryCatalog:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.catalog = self.load_catalog()

    def load_catalog(self):
        catalog = {}
        try:
            workbook = openpyxl.load_workbook(self.excel_file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                title, genre, volumes = row
                book = Book(title, genre, volumes)
                catalog[title] = book
        except Exception as e:
            print(f"Error loading catalog: {e}")
        return catalog

    def add_book_to_catalog(self, book):
        # Use the title as a unique identifier
        self.catalog[book.title] = book
        self.save_catalog()

    def save_catalog(self):
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Title", "Genre", "Volumes"])
            for book in self.catalog.values():
                sheet.append([book.title, book.genre, book.volumes])
            workbook.save(self.excel_file)
        except Exception as e:
            print(f"Error saving catalog: {e}")

    def display_catalog(self):
        for book in self.catalog.values():
            print(book.get_info())

    def sort_catalog_by_title(self):
        sorted_catalog = sorted(self.catalog.values(), key=lambda x: x.title)
        self.catalog = {book.title: book for book in sorted_catalog}
        self.save_catalog()

    def sort_catalog_by_genre(self):
        sorted_catalog = sorted(self.catalog.values(), key=lambda x: x.genre)
        self.catalog = {book.title: book for book in sorted_catalog}
        self.save_catalog()

    def sort_catalog_by_volumes(self):
        sorted_catalog = sorted(self.catalog.values(), key=lambda x: x.volumes)
        self.catalog = {book.title: book for book in sorted_catalog}
        self.save_catalog()

    def delete_book(self, title):
        if title in self.catalog:
            del self.catalog[title]
            self.save_catalog()
            return True
        else:
            return False
